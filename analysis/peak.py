import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math
from analysis.sheet_manager import SheetManager
import xlwt
import openpyxl
from pathlib import Path
from analysis.file_utils import get_vial_names, get_all_txt_files


def pick_peak(RT_dict: dict, sheet_manager: SheetManager, data_dir: Path):
    work_sheet = sheet_manager.load_peak_id_sheet()
    df = sheet_manager.load_john_code_data_frames()
    # Preparing list of strings called Vial_names where each element is the name of a sample
    txt_files = get_all_txt_files(data_dir)
    Vial_names = get_vial_names(txt_files)
    # Create list of analytes to make it easy to call dictionary in future
    analytes = RT_dict.keys()

    # This block goes through the dataframe from the John_Code sheet and picks out sample names and relevant peaks.
    txt_file_headers = ['R.time', 'I.time',
                        'F.time', 'Area', 'Height', 'Peak_ID']

    for i in range(len(df)):
        # picks out each row in dataframe and truncates - would be more elegant to not have to use dataframe but whatever
        row_calling = df.loc[i, :]
        vals = row_calling.values.tolist()
        vals = vals[0:6]

        y = type(vals[0])

        if vals[0] in Vial_names:
            sample_of_interest = 'yes'

            # picks out sample names and adds to worksheet
            print(vals[0])
            heading = [vals[0], vals[0], vals[0],
                       vals[0], vals[0], vals[0]]
            print(heading)
            work_sheet.append(heading)
            work_sheet.append(txt_file_headers)

        elif vals[0] != 'Peak#' and y == str:
            sample_of_interest = 'no'

        # picks out peaks
        x = type(vals[1])
        if x == int:  # This if statement is for the rare occasion when the retention time of an analyte is an integer
            vals[1] = float(vals[1])
            x = type(vals[1])

        if x == float and sample_of_interest == 'yes':
            for a in analytes:
                upper_tolerance = RT_dict[a] + 0.1
                lower_tolerance = RT_dict[a] - 0.1
                if vals[1] < upper_tolerance and vals[1] > lower_tolerance:
                    vals.append(a)
                    work_sheet.append(vals[1:])


def bala():
    wb = openpyxl.load_workbook(filename)
    sheet2 = ''
    sheet2 = wb['Peak_ID']
    sheet3 = ''

    sheet3 = wb.create_sheet('Quantification w IS,ES')

    ID = 'blank'
    Area = 'blank'
    sheet3_list = []

# iterates through a semi-arbitrary number of rows and only reports values, still a tuple
    for row in sheet2.iter_rows(min_row=1, max_row=5000, min_col=1, max_col=10, values_only=True):

        if row[5] == ID:  # If there is a duplicate label for a single peak, pick the peak with the larger area
            if row[3] > Area:
                sheet3_list.pop()
                sheet3_list.append(row)
            else:
                continue

        else:
            sheet3_list.append(row)

        if row[3] == None:  # breaks out of for loop once all text file data is iterated through
            break

        else:
            ID = row[5]
            Area = row[3]

    for row in sheet3_list:
        sheet3.append(row)

# Notice that this code may not work as intended if there is a duplicate in the first row, however, since the first row will always be a sample name, it is fine.

    wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    sheet_read = ''
    sheet_read = wb['Peak_ID']
    sheet3 = ''
    sheet3 = wb.create_sheet('Quantification w IS,ES-full')

    sample_header = []
    data_header = ['R.time', 'I.time', 'F.time', 'Area', 'Height', 'Peak_ID']

# iterates through a semi-arbitrary number of rows and only reports values, still a tuple
    for row in sheet_read.iter_rows(min_row=1, max_row=sheet_read.max_row, min_col=1, max_col=10, values_only=True):
        # This is a bit clunky, but it ensures the order of the samples in the Peak_ID tab is maintained
        if row[0] == row[1] and row[1] == row[2]:
            sample = row[0]
            sample_header = [sample, sample, sample, sample, sample, sample]
            sheet3.append(sample_header)
            sheet3.append(data_header)
            for a in analytes:
                blank_row = [0, 0, 0, 0, 0, a]
                sheet3.append(blank_row)

    print(sheet3.max_row)


#wb = openpyxl.load_workbook('GCData-JGI_ACRs.xlsx')
    sheet2 = ''
    sheet2 = wb['Quantification w IS,ES']
    sheet3 = ''
    sheet3 = wb['Quantification w IS,ES-full']

    Ass_Fat = True
    last_count = 1
    while Ass_Fat:

        counter = 0

        for row2, row3 in zip(sheet2.iter_rows(min_row=1, max_row=sheet3.max_row, min_col=1, max_col=10, values_only=True), sheet3.iter_rows(min_row=1, max_row=sheet3.max_row, min_col=1, max_col=10, values_only=True)):
            #print(row1[0], row2[0])
            counter = counter + 1
            # print(counter)
            if counter < last_count:
                continue
            elif row2[5] == row3[5]:
                last_count = counter
                print(last_count)
            elif row2[5] != row3[5]:
                # print(row3[5])
                sheet2.insert_rows(counter)
                sheet2.cell(row=counter, column=6).value = row3[5]
                sheet2.cell(row=counter, column=5).value = 0.000000001
                #last_count = counter
                break

            if counter == sheet3.max_row:
                Ass_Fat = False
                break
            else:
                continue


#wb.remove_sheet('Quantification w IS,ES-full')
    wb.save(filename)

# useless
    for i, j in zip([1, 2, 3], [3, 2, 1]):
        print(i, j)
