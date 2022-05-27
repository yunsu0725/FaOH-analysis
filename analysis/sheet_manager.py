from email import header
from pathlib import Path
import openpyxl
import pandas as pd
from typing import Dict, List

from analysis.data_point import DataPoint, pad_missing_chain, parse_data_point_from_quant_sheet_row


class SheetManager:
    def __init__(self, file_path: Path) -> None:
        self.file_path = file_path
        self.preprocess_key = 'Preprocessed Data'
        self.quant_key = 'Quantification w IS,ES'
        self.external_standard_key = 'EXT_STD'
        self.concentration_key = 'Corrected Concentration'
        # TODO: remove
        self.peak_id_key = 'Peak_ID'
        self.quant_full_key = 'Quantification w IS,ES-full'
        self.wb = openpyxl.Workbook()

    def create(self):
        self.wb.create_sheet(title=self.preprocess_key)
        # self.wb.create_sheet(title=self.peak_id_key)
        self.wb.create_sheet(title=self.quant_key)
        # self.wb.create_sheet(title=self.quant_full_key)
        self.wb.create_sheet(title=self.external_standard_key)
        self.wb.create_sheet(title=self.concentration_key)
        del self.wb['Sheet']

    def load_raw_sheet(self):
        return self.wb[self.preprocess_key]

    def load_raw_data_frames(self):
        df = pd.read_excel(
            self.file_path, sheet_name=self.preprocess_key, header=None
        )
        return df

    def load_quant_sheet_data_frame(self, use_cols=None) -> pd.DataFrame:
        df = pd.read_excel(
            self.file_path, sheet_name=self.quant_key, header=None, usecols=use_cols
        )
        return df

    def load_concentration_sheet(self):
        return self.wb[self.concentration_key]

    # TODO: deprecate
    def load_peak_id_sheet(self):
        return self.wb[self.peak_id_key]

    def load_external_standard_sheet(self):
        return self.wb[self.external_standard_key]

    def load_quant_sheet(self):
        return self.wb[self.quant_key]

    # TODO: deprecate
    def load_quant_full_sheet(self):
        return self.wb[self.quant_full_key]

    # TODO: deprecate
    def drop_tmp_sheets(self):
        # it is a tmp workaround to drop the quant full sheet
        # we shd remove that sheet from the logic itself for long term if have time
        del self.wb[self.quant_full_key]
        self.save_workbook()

    def write_raw_data_points(self, dp_dict: Dict[str, List[DataPoint]], vials: List[str]):
        sheet = self.load_raw_sheet()
        headers = DataPoint.get_raw_data_sheet_header()
        for vial in vials:
            try:
                data_points = dp_dict[vial]
                sheet.append([vial])
                sheet.append(headers)
                for dp in data_points:
                    sheet.append(dp.get_raw_data_row())
            except KeyError:
                # TODO: error logger
                pass

    def write_quantification_sheet(
        self,
        dp_dict: Dict[str, List[DataPoint]],
        vials: List[str],
        analytes: List[str],
    ):
        sheet = self.load_quant_sheet()
        headers = DataPoint.get_quantification_sheet_header()
        for vial in vials:
            try:
                data_points = dp_dict[vial]
                sheet.append([vial for _ in range(len(headers))])
                sheet.append(headers)
                pad_dps = pad_missing_chain(data_points, analytes)
                for dp in pad_dps:
                    sheet.append(dp.get_quantification_sheet_row())
            except KeyError:
                # TODO: error logger
                pass

    def save_workbook(self):
        self.wb.save(self.file_path)

    def load_data_points_from_quant_sheet(self, vial_names: List[str]) -> Dict[str, List[DataPoint]]:
        """construct the data points from the data frame

        Args:
            vial_names (List[str]): all vial names

        Returns:
            Dict[str, List[DataPoint]]: the data points dict, the key is sth like FaOH-10
        """
        df = self.load_quant_sheet_data_frame()
        df_list = df.values.tolist()

        vial_set = set(vial_names)
        cur_vial = None
        dp_list = list()
        res = dict()

        for row in df_list:
            first_elem = row[0]
            if first_elem in vial_set:
                # ex row: ['FaOH-10', nan, nan, nan, nan, nan]
                if cur_vial is not None:
                    res[cur_vial] = dp_list

                dp_list = list()
                cur_vial = first_elem
                continue

            if row == DataPoint.get_quantification_sheet_header():
                # ignore the header row
                # shd be sth like: ['R.time', 'I.time', 'F.time', 'Area', 'Height', 'Peak_ID']
                continue

            dp = parse_data_point_from_quant_sheet_row(row)
            dp_list.append(dp)

        # the last vial in the sheet
        if cur_vial is not None:
            res[cur_vial] = dp_list

        return res
